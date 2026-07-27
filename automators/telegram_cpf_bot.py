import asyncio
import re
import os
import io
from PIL import Image
import ddddocr
from openpyxl import load_workbook
from playwright.async_api import async_playwright

import sys

# 导入中心化配置
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import SCRIPT_TEMPLATE

# 初始化 OCR，关闭广告输出
ocr = ddddocr.DdddOcr(show_ad=False)

EXCEL_PATH = SCRIPT_TEMPLATE
USER_DATA_DIR = os.path.join(os.getcwd(), "telegram_session")
CHAT_NAME = "Skynet Robot (Privado)"

def solve_math_captcha(image_bytes):
    """处理并识别数学验证码截图"""
    try:
        # 0. 借助目标检测找运算符 (加号高度大，减号高度极小)
        detected_op = None
        try:
            det = ddddocr.DdddOcr(det=True, show_ad=False)
            poses = det.detection(image_bytes)
            
            # 首先检查 ddddocr 是否直接框出了减号 (宽远大于高)
            for b in poses:
                bw, bh = b[2] - b[0], b[3] - b[1]
                if bw > 10 and bh > 0 and bw > bh * 2.5:
                    detected_op = '-'
                    break
            
            # 如果没直接检测到减号，就去数字中间的缝隙里找
            if not detected_op:
                # 过滤出高度大于 20 的框，这些才是数字！(减号高度很小)
                digit_boxes = [b for b in poses if b[3] - b[1] > 20]
                digit_boxes = sorted(digit_boxes, key=lambda b: b[0])
                
                if len(digit_boxes) >= 2:
                    # 有时候识别出了加号（也是个大框），为了保险，我们只看最左边和最右边的两个数字框
                    box1, box2 = digit_boxes[0], digit_boxes[-1]
                    img = Image.open(io.BytesIO(image_bytes)).convert('RGB')
                    pixels = img.load()
                    for x in range(img.width):
                        for y in range(img.height):
                            r, g, b = pixels[x, y]
                            if r > 160 and g > 160 and b > 160:
                                pixels[x, y] = (255, 255, 255)
                            else:
                                pixels[x, y] = (0, 0, 0)
                    img = img.convert('L')
                    pixels = img.load()
                            
                    # 通过两个数字的中心点来定位符号中心
                    c1 = (box1[0] + box1[2]) / 2
                    c2 = (box2[0] + box2[2]) / 2
                    gap_center = (c1 + c2) / 2
                    
                    # 使用连通域分析（BFS）来寻找运算符，彻底解决数字靠得太近导致的裁剪重叠问题
                    D = c2 - c1
                    w_img, h_img = img.size
                    
                    visited = [[False]*h_img for _ in range(w_img)]
                    components = []
                    
                    for x in range(w_img):
                        for y in range(h_img):
                            if pixels[x, y] == 255 and not visited[x][y]:
                                comp_pixels = []
                                q = [(x, y)]
                                visited[x][y] = True
                                while q:
                                    cx, cy = q.pop(0)
                                    comp_pixels.append((cx, cy))
                                    for dx in [-1, 0, 1]:
                                        for dy in [-1, 0, 1]:
                                            if dx == 0 and dy == 0: continue
                                            nx, ny = cx + dx, cy + dy
                                            if 0 <= nx < w_img and 0 <= ny < h_img:
                                                if pixels[nx, ny] == 255 and not visited[nx][ny]:
                                                    visited[nx][ny] = True
                                                    q.append((nx, ny))
                                
                                if len(comp_pixels) >= 4: # 过滤极小噪点
                                    xs = [p[0] for p in comp_pixels]
                                    ys = [p[1] for p in comp_pixels]
                                    cx_comp = (min(xs) + max(xs)) / 2
                                    components.append({
                                        'min_x': min(xs), 'max_x': max(xs),
                                        'min_y': min(ys), 'max_y': max(ys),
                                        'cx': cx_comp
                                    })
                                    
                    # 找到所有落在 gap_center 附近（安全区）的连通域
                    safe_dist = max(5, D / 3.5)
                    op_comps = [c for c in components if abs(c['cx'] - gap_center) <= safe_dist]
                    
                    if op_comps:
                        # 合并这些连通域的边界（以防加号或等号断裂成几块）
                        min_x = min([c['min_x'] for c in op_comps])
                        max_x = max([c['max_x'] for c in op_comps])
                        min_y = min([c['min_y'] for c in op_comps])
                        max_y = max([c['max_y'] for c in op_comps])
                        
                        op_w = max_x - min_x + 1
                        op_h = max_y - min_y + 1
                        
                        # 如果高度极低，毫无疑问是一根横杠（减号）
                        if op_h <= 8 or op_w > op_h * 1.5:
                            detected_op = '-'
                        else:
                            # 区分加号(+)和乘号(x或*)：
                            # 加号只有横竖两笔，所以左上、右上、左下、右下四个“角落”绝对是空无一物的。
                            # 乘号带有交叉斜线，它的四个“角落”一定会扫到像素。
                            corner_pixels = 0
                            cw = max(1, op_w // 4)
                            ch = max(1, op_h // 4)
                            
                            for y in range(min_y, min_y + ch):
                                for x in range(min_x, min_x + cw):
                                    if pixels[x, y] == 255: corner_pixels += 1
                                for x in range(max_x - cw + 1, max_x + 1):
                                    if pixels[x, y] == 255: corner_pixels += 1
                            for y in range(max_y - ch + 1, max_y + 1):
                                for x in range(min_x, min_x + cw):
                                    if pixels[x, y] == 255: corner_pixels += 1
                                for x in range(max_x - cw + 1, max_x + 1):
                                    if pixels[x, y] == 255: corner_pixels += 1
                                    
                            if corner_pixels >= 12:
                                detected_op = '*'
                            else:
                                detected_op = '+'
        except Exception as e:
            print(f"[!] 物理分析运算符出错: {e}")
            
        # 1. 加载并转为 RGB 模式进行色彩过滤
        image = Image.open(io.BytesIO(image_bytes)).convert('RGB')
        
        # 2. RGB 二值化：严格过滤掉带色彩的干扰线，仅保留纯白/灰白的高亮文字
        pixels = image.load()
        width, height = image.size
        for x in range(width):
            for y in range(height):
                r, g, b = pixels[x, y]
                if r > 160 and g > 160 and b > 160:
                    pixels[x, y] = (255, 255, 255)
                else:
                    pixels[x, y] = (0, 0, 0)
                    
        image = image.convert('L')
                    
        # 3. 将净化后的图片转换为 bytes
        img_byte_arr = io.BytesIO()
        image.save(img_byte_arr, format='PNG')
        clean_bytes = img_byte_arr.getvalue()
        
        # 4. OCR 识别
        res = ocr.classification(clean_bytes)
        print(f"[*] OCR 原始识别结果: {res}")
        
        # 尝试提取明确的运算符
        res_raw = res.lower().replace('x', '*').replace(' ', '')
        
        operator = None
        # 0. 乘号特判防覆盖：OCR 对 'x' 的识别极准，若明确包含乘号则直接信任 OCR，防止被物理规则误判为减号
        if '*' in res_raw:
            operator = '*'
        # 1. 物理测量出的减号最准
        elif detected_op == '-':
            operator = '-'
        # 2. 物理测量出的乘号极准（因为加号角落必定为空）
        elif detected_op == '*':
            operator = '*'
        # 3. 物理测量出的加号很准
        elif detected_op == '+':
            operator = '+'
        # 4. 如果物理测量完全没生效（比如数字贴得太近被识别成了一个框），则完全信任 OCR
        else:
            if '*' in res_raw:
                operator = '*'
            elif '+' in res_raw:
                operator = '+'
            elif '-' in res_raw:
                operator = '-'
                
        # 增加基于用户经验的修正规则：如果 OCR 结果像 847，中间通常是符号（例如 + 被错认为 4）
        res_clean = re.sub(r'[^\d]', '', res)
        if not operator and len(res_clean) >= 3:
            if res_clean[1] == '4':
                operator = '+'
                print("[*] 智能修正: 推测中间的 '4' 实际上是 '+' 号")
            
        if operator:
            print(f"[*] 确认运算符为: {operator}")
            
        
        # 5. 解析操作数
        n1, n2 = None, None
        try:
            if 'digit_boxes' in locals() and len(digit_boxes) >= 2:
                def get_digit(box):
                    x1, y1, x2, y2 = box
                    x1, y1 = max(0, x1-10), max(0, y1-10)
                    x2, y2 = min(width, x2+10), min(height, y2+10)
                    crop = image.crop((x1, y1, x2, y2))
                    arr = io.BytesIO()
                    crop.save(arr, format='PNG')
                    txt = ocr.classification(arr.getvalue())
                    d = re.sub(r'[^\d]', '', txt)
                    return int(d[0]) if d else None
                    
                n1 = get_digit(digit_boxes[0])
                n2 = get_digit(digit_boxes[1])
                if n1 is not None and n2 is not None:
                    print(f"[*] 物理裁剪提取操作数成功: n1={n1}, n2={n2}")
        except Exception as e:
            print(f"[!] 物理裁剪提取数字出错: {e}")
            
        # 如果物理裁剪失败，退回到全局字符串解析
        if n1 is None or n2 is None:
            if len(res_clean) == 2:
                n1, n2 = int(res_clean[0]), int(res_clean[1])
                print(f"[*] 全局字符串提取操作数: n1={n1}, n2={n2}")
            elif len(res_clean) >= 3:
                # 遵循用户经验：“中间是不会有数字的”
                # 如果出现3个以上的数字串，真正的操作数必定在两端，中间的是被误识别的符号
                n1, n2 = int(res_clean[0]), int(res_clean[2])
                print(f"[*] 优化提取操作数 (忽略中间的干扰数字): n1={n1}, n2={n2}")
        
        ordered_ans = []
        if n1 is not None and n2 is not None:
            ans_plus = n1 + n2
            ans_mul = n1 * n2
            # 只有当 n1 >= n2 时，相减才为非负数（验证码选项中不会有负数，因此 n1 < n2 时绝不可能是减法）
            ans_minus = n1 - n2 if n1 >= n2 else None
            
            def add_ans(a):
                if a is not None:
                    ordered_ans.append(a)
            
            if operator == '+':
                add_ans(ans_plus)
                add_ans(ans_minus)
                add_ans(ans_mul)
            elif operator == '-':
                add_ans(ans_minus)
                add_ans(ans_plus)
                add_ans(ans_mul)
            elif operator == '*':
                add_ans(ans_mul)
                add_ans(ans_plus)
                add_ans(ans_minus)
            else:
                # 盲猜优先级：加 > 减 > 乘
                add_ans(ans_plus)
                add_ans(ans_minus)
                add_ans(ans_mul)
            
            # 为了最大化容错，如果提取出来的数字不止两个，全部两两组合
            res_clean = re.sub(r'[^\d]', '', res)
            if len(res_clean) > 2:
                digits = [int(d) for d in res_clean]
                for i in range(len(digits)):
                    for j in range(i+1, len(digits)):
                        d1, d2 = digits[i], digits[j]
                        ordered_ans.append(d1 + d2)
                        if d1 >= d2:
                            ordered_ans.append(d1 - d2)
                        ordered_ans.append(d1 * d2)
                    
        certain_ans_str = None
        if n1 is not None and n2 is not None and operator is not None:
            if operator == '+':
                certain_ans_str = str(n1 + n2)
            elif operator == '-':
                if n1 >= n2:
                    certain_ans_str = str(n1 - n2)
            elif operator == '*':
                certain_ans_str = str(n1 * n2)

        # 去重并保持优先级顺序
        seen = set()
        possible_ans_str = []
        for a in ordered_ans:
            if str(a) not in seen:
                seen.add(str(a))
                possible_ans_str.append(str(a))
                
        if certain_ans_str:
            print(f"[*] 确定的答案为: {certain_ans_str}")
        print(f"[*] 从OCR结果 '{res}' 猜测的可能答案 (按优先级): {possible_ans_str}")
        return certain_ans_str, possible_ans_str
    except Exception as e:
        print(f"[!] 验证码处理出错: {e}")
        return None, []

async def run_cpf_query(excel_path=EXCEL_PATH, user_data_dir=USER_DATA_DIR, headless=False, progress_callback=None):
    def log(msg):
        print(msg)
        if progress_callback:
            progress_callback(msg)
            
    log(f"[*] 加载 Excel 文件: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    async with async_playwright() as p:
        log("[*] 启动浏览器...")
        context = await p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            channel="chrome", # 使用您电脑上自带的 Google Chrome
            headless=headless,
            viewport={'width': 1280, 'height': 800}
        )
        
        page = await context.new_page()
        log(f"[*] 正在打开 Telegram Web...")
        await page.goto('https://web.telegram.org/k/', wait_until='domcontentloaded', timeout=120000)
        
        log(f"[*] 正在左侧聊天列表中寻找 '{CHAT_NAME}' ...")
        try:
            chat_locator = page.locator(f'text="{CHAT_NAME}"').first
            await chat_locator.wait_for(state='visible', timeout=60000)
            await chat_locator.click()
        except Exception as e:
            log(f"[!] 找不到对应的聊天对象，请检查是否在左侧列表中。错误信息: {e}")
            await context.close()
            return
            
        input_selector = 'div.input-message-input'
        await page.wait_for_selector(input_selector, timeout=30000)
        log("[*] 聊天界面加载完毕！开始处理表格数据...")
        
        last_successful_reply_text = ""
        last_processed_cpf = ""
        
        async def process_cpf_row(r_idx, is_retry=False):
            nonlocal last_successful_reply_text, last_processed_cpf
            cpf_cell = ws.cell(row=r_idx, column=4).value # D列
            
            if not cpf_cell:
                if not is_retry:
                    log(f"[*] 第 {r_idx} 行遇到空号码，此遍处理完成！")
                return False
                
            status_cell = ws.cell(row=r_idx, column=5).value # E列
            status_str = str(status_cell).strip() if status_cell else ""
            
            if not is_retry:
                if status_str != "":
                    log(f"[*] 第 {r_idx} 行已有结果 ({status_str})，跳过...")
                    return True
            else:
                if status_str not in ["遇到验证码且未能通过", "查询超时", "提取失败"]:
                    return True

            cpf_text = str(cpf_cell).strip()
            if is_retry:
                log(f"\n[{r_idx}] (重试) 正在重新查询之前失败的 CPF: {cpf_text}")
            else:
                log(f"\n[{r_idx}] 正在查询 CPF: {cpf_text}")
            
            pre_query_text = ""
            pre_in_msg_count = 0
            
            pre_query_msg_id = None
            
            async def close_popup_if_any():
                try:
                    clicked = await page.evaluate('''() => {
                        let text = document.body.innerText || "";
                        if (text.includes("resolvido") || text.includes("já pode continuar")) {
                            let btns = Array.from(document.querySelectorAll('button, .btn, [role="button"]'));
                            let okBtn = btns.find(b => b.innerText && b.innerText.trim().toUpperCase() === 'OK');
                            if (okBtn) {
                                okBtn.click();
                                return true;
                            }
                        }
                        return false;
                    }''')
                    if clicked:
                        log("[*] 已成功侦测并强制关闭了 OK 弹窗！")
                        await asyncio.sleep(0.5)
                        return True
                except Exception as e:
                    log(f"[!] 关闭弹窗时出错: {e}")
                return False

            async def send_query():
                nonlocal pre_query_text, pre_in_msg_count, pre_query_msg_id
                await close_popup_if_any()
                last_in_msgs = await page.query_selector_all('div.bubble.is-in .message')
                if last_in_msgs:
                    pre_query_text = await last_in_msgs[-1].inner_text()
                    bubble = await last_in_msgs[-1].evaluate_handle('el => el.closest(".bubble")')
                    pre_query_msg_id = await bubble.get_attribute('data-mid') if bubble else None
                else:
                    pre_query_msg_id = None
                pre_in_msg_count = len(last_in_msgs)
                    
                msg_count = len(await page.query_selector_all('div.message'))
                await page.fill(input_selector, cpf_text)
                await page.press(input_selector, 'Enter')
                for _ in range(25):
                    await asyncio.sleep(0.2)
                    if len(await page.query_selector_all('div.message')) > msg_count:
                        break

            await send_query()
            
            reply_text = ""
            attempts = 0
            last_handled_captcha = None
            verification_failed_id = None
            last_handled_leftover_id = None
            is_timeout = False
            
            while True:
                if attempts >= 100:
                    is_timeout = True
                    break
                await close_popup_if_any()
                await asyncio.sleep(0.2)
                message_elements = await page.query_selector_all('div.bubble.is-in .message')
                if not message_elements:
                    attempts += 1
                    continue
                    
                last_message = message_elements[-1]
                reply_text = await last_message.inner_text()
                
                bubble_el = await last_message.evaluate_handle('el => el.closest(".bubble")')
                msg_id = await bubble_el.get_attribute('data-mid') if bubble_el else None
                
                # 判断是否是新消息：基于 msg_id 能够完全无视虚拟滚动导致的 DOM 数量变化
                if pre_query_msg_id is None:
                    is_new_reply = True
                elif msg_id != pre_query_msg_id:
                    is_new_reply = True
                elif reply_text != pre_query_text:
                    is_new_reply = True
                else:
                    is_new_reply = False
                
                if not is_new_reply:
                    attempts += 1
                    continue
                
                captcha_id = msg_id if msg_id else reply_text
                
                text_lower = reply_text.lower()
                
                # ------ 防残留消息机制 ------
                if reply_text == last_successful_reply_text and cpf_text != last_processed_cpf:
                    if msg_id != last_handled_leftover_id:
                        last_handled_leftover_id = msg_id
                        log(f"[!] 警告：检测到上一个单号 ({last_processed_cpf}) 的延迟残留回复，自动忽略...")
                    attempts += 1
                    continue
                
                if "não encontrado" in text_lower or "nao encontrado" in text_lower or "inválido" in text_lower or "invalido" in text_lower or "utilize o comando" in text_lower or "seguido dos" in text_lower or "dígitos do cpf" in text_lower:
                    break

                if "nome" in text_lower:
                    break
                        
                if "muitas requisições" in text_lower or "suspenso" in text_lower or "captcha" in text_lower or "errado" in text_lower:
                    bubble = await last_message.evaluate_handle('el => el.closest(".bubble")')
                    
                    img_el = None
                    img_src = ""
                    if bubble:
                        img_el = await bubble.query_selector('.media-photo')
                        if not img_el:
                            img_el = await bubble.query_selector('img')
                        if img_el:
                            img_src = await img_el.get_attribute('src') or ""
                            
                    captcha_id = f"{msg_id}_{img_src}" if msg_id else f"{reply_text}_{img_src}"
                    
                    if captcha_id == last_handled_captcha:
                        attempts += 1
                        continue
                        
                    log(f"[!] 检测到验证码！开始尝试识别...")
                    
                    await asyncio.sleep(0.5)
                    captcha_id = f"{msg_id}_{img_src}" if msg_id else f"{reply_text}_{img_src}"

                    if not img_el:
                        buttons_check = await bubble.query_selector_all('button') if bubble else []
                        if not buttons_check:
                            if captcha_id != verification_failed_id:
                                verification_failed_id = captcha_id
                                log(f"[!] 消息含关键词但无图片无按钮: {reply_text[:80]!r}")
                                match_min = re.search(r'(\d+)\s*minutos', reply_text.lower())
                                if match_min:
                                    if not getattr(send_query, "wait_min_attempts", False):
                                        send_query.wait_min_attempts = 1
                                        log(f"[*] 收到验证码过期/等待提示: {reply_text}。先尝试立即重新发送一次，看是否能刷出新验证码...")
                                        await send_query()
                                        
                                        log("[*] 重发后扫描屏幕，确认是否有遗漏的验证码...")
                                        await asyncio.sleep(2)
                                        recent_bubbles = await page.query_selector_all('div.bubble.is-in')
                                        found_captcha = False
                                        for b_idx in range(len(recent_bubbles)-1, max(-1, len(recent_bubbles)-6), -1):
                                            b = recent_bubbles[b_idx]
                                            img_el = await b.query_selector('.media-photo')
                                            if not img_el:
                                                img_el = await b.query_selector('img')
                                            btns = await b.query_selector_all('button')
                                            if img_el and btns:
                                                log("[*] 找到了遗漏的验证码！开始尝试识别...")
                                                await asyncio.sleep(0.5)
                                                try:
                                                    image_bytes = await img_el.screenshot()
                                                    certain_ans, possible_answers = solve_math_captcha(image_bytes)
                                                    btn_texts = []
                                                    for btn in btns:
                                                        txt = (await btn.inner_text()).strip()
                                                        btn_texts.append((btn, txt))
                                                    
                                                    target_btn, target_txt = None, None
                                                    if certain_ans:
                                                        for btn, txt in btn_texts:
                                                            if txt == certain_ans:
                                                                target_btn, target_txt = btn, txt
                                                                break
                                                    if not target_btn and possible_answers:
                                                        matched = [ (btn, txt) for btn, txt in btn_texts if txt in possible_answers ]
                                                        if len(matched) == 1:
                                                            target_btn, target_txt = matched[0]
                                                            
                                                    if target_btn:
                                                        log(f"[*] 自动点击答案 '{target_txt}'...")
                                                        try:
                                                            await target_btn.click(force=True, delay=100)
                                                        except Exception as e:
                                                            await target_btn.evaluate('b => b.click()')
                                                        log("[*] 验证码已解答，正在等待确认弹窗...")
                                                        for _ in range(15):
                                                            await asyncio.sleep(0.2)
                                                            if await close_popup_if_any():
                                                                break
                                                        log("[*] 重发因验证码未解答而错过的 CPF...")
                                                        await send_query()
                                                        found_captcha = True
                                                    else:
                                                        log("[!] 遗漏的验证码存在歧义或无法识别，请人工介入：请手动点击正确答案！(等待60秒)...")
                                                        manual_solved = False
                                                        for _ in range(300):
                                                            await asyncio.sleep(0.2)
                                                            if await close_popup_if_any():
                                                                manual_solved = True
                                                                break
                                                        if manual_solved:
                                                            log("[*] 检测到人工已解决验证码！重发错过的 CPF...")
                                                            await send_query()
                                                            found_captcha = True
                                                        else:
                                                            log("[!] 等待人工解决超时！根据要求重新扫描并尝试强行盲猜一个答案...")
                                                            recent_bubbles = await page.query_selector_all('div.bubble.is-in')
                                                            guessed_btn = None
                                                            guessed_txt = None
                                                            for gb_idx in range(len(recent_bubbles)-1, max(-1, len(recent_bubbles)-6), -1):
                                                                gb = recent_bubbles[gb_idx]
                                                                gimg_el = await gb.query_selector('.media-photo')
                                                                if not gimg_el:
                                                                    gimg_el = await gb.query_selector('img')
                                                                gbtns = await gb.query_selector_all('button')
                                                                if gimg_el and gbtns:
                                                                    try:
                                                                        gimage_bytes = await gimg_el.screenshot()
                                                                        _, gpos = solve_math_captcha(gimage_bytes)
                                                                        gbtn_texts = [(btn, (await btn.inner_text()).strip()) for btn in gbtns]
                                                                        if gpos:
                                                                            for btn, txt in gbtn_texts:
                                                                                if txt in gpos:
                                                                                    guessed_btn, guessed_txt = btn, txt
                                                                                    break
                                                                        if not guessed_btn and gbtn_texts:
                                                                            guessed_btn, guessed_txt = gbtn_texts[0]
                                                                    except Exception as ge:
                                                                        log(f"[!] 盲猜出错: {ge}")
                                                                        if gbtns:
                                                                            guessed_btn = gbtns[0]
                                                                    break
                                                            
                                                            if guessed_btn:
                                                                log(f"[*] 强行盲猜答案 '{guessed_txt}' 并点击...")
                                                                try:
                                                                    await guessed_btn.click(force=True, delay=100)
                                                                except:
                                                                    await guessed_btn.evaluate('b => b.click()')
                                                                await asyncio.sleep(1)
                                                                await close_popup_if_any()
                                                                log("[*] 已盲猜，重发 CPF...")
                                                                await send_query()
                                                                found_captcha = True
                                                except Exception as e:
                                                    log(f"[!] 处理遗漏验证码时出错: {e}")
                                                break
                                                
                                        if not found_captcha:
                                            log("[*] 未发现验证码，等待后续处理 (若再次收到警告则真正休眠)...")
                                            
                                        attempts = 0
                                    else:
                                        wait_mins = int(match_min.group(1))
                                        log(f"[*] 再次收到等待提示，确实需要等待 {wait_mins} 分钟。开始休眠...")
                                        await asyncio.sleep(wait_mins * 60)
                                        send_query.wait_min_attempts = 0
                                        log(f"[*] 休眠结束，重新发送 CPF: {cpf_text} ...")
                                        await send_query()
                                        attempts = 0
                                elif "muitas requisições" in reply_text.lower() or "suspenso" in reply_text.lower():
                                    log(f"[*] 判定为频率限制或封禁，5 分钟后重新发送 CPF: {cpf_text} ...")
                                    await asyncio.sleep(300)
                                    await send_query()
                                    attempts = 0
                                    log(f"[*] 已重新发送，继续等待回复...")
                                else:
                                    log("[*] 判定为普通警告提示 (如验证码过期)，忽略此消息并继续等待...")
                            else:
                                attempts += 1
                            continue

                    if not bubble:
                        log("[!] 找不到验证码的 bubble 容器")
                        break
                        
                    if not img_el:
                        img_el = last_message
                        
                    target_btn = None
                    target_txt = None
                    possible_answers = None
                    image_bytes = None
                    
                    for scan_attempt in range(4):
                        if scan_attempt > 0:
                            log(f"[!] 第 {scan_attempt} 次尝试未得到确信结果，等待 5 秒后重试扫描...")
                            await asyncio.sleep(5)
                            # 重新获取图片元素，以防 DOM 刷新
                            if bubble:
                                img_el = await bubble.query_selector('.media-photo')
                                if not img_el:
                                    img_el = await bubble.query_selector('img')
                            if not img_el:
                                img_el = last_message
                                
                        try:
                            image_bytes = await img_el.screenshot()
                            certain_ans, possible_answers = solve_math_captcha(image_bytes)
                            
                            buttons = await bubble.query_selector_all('button')
                            wait_sec = 0
                            while not buttons and wait_sec < 5:
                                log(f"[!] 暂时未找到按钮，等待中 ({wait_sec}s/5s)...")
                                await asyncio.sleep(1)
                                wait_sec += 1
                                buttons = await bubble.query_selector_all('button')
                                
                            log(f"[*] 第 {scan_attempt+1} 次尝试提取到 {len(buttons)} 个按钮")
                            btn_texts = []
                            for btn in buttons:
                                txt = (await btn.inner_text()).strip()
                                btn_texts.append((btn, txt))
                                
                            if certain_ans:
                                for btn, txt in btn_texts:
                                    if txt == certain_ans:
                                        target_btn, target_txt = btn, txt
                                        log(f"[*] 确信答案 '{txt}' 存在于选项中，直接使用。")
                                        break
                                        
                            if not target_btn and possible_answers:
                                matched_guesses = []
                                for ans_str in possible_answers:
                                    for btn, txt in btn_texts:
                                        if txt == ans_str:
                                            matched_guesses.append((btn, txt))
                                            break
                                if len(matched_guesses) == 1:
                                    target_btn, target_txt = matched_guesses[0]
                                    log(f"[*] 虽然没有确信答案，但在选项中只有 '{target_txt}' 这一个可能，推断为唯一解...")
                                elif len(matched_guesses) > 1:
                                    log(f"[!] 选项中存在多个可能答案 {[m[1] for m in matched_guesses]}，存在歧义，拒绝瞎猜。")
                                    
                        except Exception as e:
                            log(f"[!] 获取验证码图片或按钮失败 (可能DOM已刷新): {e}")
                            
                        if target_btn or possible_answers or certain_ans:
                            break
                    
                    if not target_btn and not possible_answers and not image_bytes:
                        attempts += 1
                        await asyncio.sleep(0.5)
                        continue

                    if target_btn:
                        log(f"[*] 正在自动点击确信答案 '{target_txt}'...")
                        try:
                            try:
                                await target_btn.click(force=True, delay=100)
                            except Exception as e:
                                await target_btn.evaluate('b => b.click()')
                        except Exception as e:
                            log(f"[!] 点击按钮失败 (可能DOM已刷新): {e}")
                            attempts += 1
                            await asyncio.sleep(0.5)
                            continue
                        
                        log("[*] 验证码已解答，正在等待并秒关确认弹窗...")
                        last_handled_captcha = captcha_id
                        for _ in range(15):
                            await asyncio.sleep(0.2)
                            if await close_popup_if_any():
                                break
                        await send_query()
                        attempts = 0
                        continue
                        
                    log(f"[!] 无法100%确认答案，或确定的答案不在选项中！")
                    if possible_answers:
                        log(f"[*] 猜测的可能答案: {possible_answers}")
                    else:
                        log(f"[*] 无法猜测出可能答案。")
                        
                    log("[!] 无法自动识别验证码，请人工介入：请手动点击正确答案！(等待60秒)...")
                    manual_solved = False
                    for _ in range(300):
                        await asyncio.sleep(0.2)
                        if await close_popup_if_any():
                            manual_solved = True
                            break
                    
                    if manual_solved:
                        log("[*] 检测到人工已解决验证码！重发 CPF...")
                        await send_query()
                        attempts = 0
                        continue
                    else:
                        log("[!] 等待人工超时，根据要求重新扫描并尝试强行盲猜一个答案...")
                        recent_bubbles = await page.query_selector_all('div.bubble.is-in')
                        guessed_btn = None
                        guessed_txt = None
                        for b_idx in range(len(recent_bubbles)-1, max(-1, len(recent_bubbles)-6), -1):
                            gb = recent_bubbles[b_idx]
                            gimg_el = await gb.query_selector('.media-photo')
                            if not gimg_el:
                                gimg_el = await gb.query_selector('img')
                            gbtns = await gb.query_selector_all('button')
                            if gimg_el and gbtns:
                                try:
                                    gimage_bytes = await gimg_el.screenshot()
                                    _, gpos = solve_math_captcha(gimage_bytes)
                                    gbtn_texts = [(btn, (await btn.inner_text()).strip()) for btn in gbtns]
                                    if gpos:
                                        for btn, txt in gbtn_texts:
                                            if txt in gpos:
                                                guessed_btn, guessed_txt = btn, txt
                                                break
                                    if not guessed_btn and gbtn_texts:
                                        guessed_btn, guessed_txt = gbtn_texts[0]
                                except:
                                    if gbtns:
                                        guessed_btn = gbtns[0]
                                break
                                
                        if guessed_btn:
                            log(f"[*] 强行盲猜答案 '{guessed_txt}' 并点击...")
                            try:
                                await guessed_btn.click(force=True, delay=100)
                            except:
                                await guessed_btn.evaluate('b => b.click()')
                            await asyncio.sleep(1)
                            await close_popup_if_any()
                            log("[*] 已盲猜，重发 CPF...")
                            await send_query()
                            attempts = 0
                            continue
                        else:
                            log("[!] 重新扫描未找到验证码，自动测试模式下直接跳过并记录失败。")
                            import time
                            with open(f"scratch/failed_captcha_{int(time.time())}.png", "wb") as f:
                                f.write(image_bytes)
                            ws.cell(row=r_idx, column=5, value="遇到验证码且未能通过")
                            try:
                                wb.save(excel_path)
                            except PermissionError:
                                pass
                            return True # Skip to next row
                

                # 注意：只要进入了 is_new_reply 为 True 的阶段，哪怕它是普通的警告，也属于新消息了
                is_captcha_msg = (
                    "muitas requisições" in text_lower
                    or "suspenso" in text_lower
                    or "captcha" in text_lower
                    or "errado" in text_lower
                )
                if not is_captcha_msg and reply_text.strip() and captcha_id != verification_failed_id:
                    verification_failed_id = captcha_id
                    log(f"[!] 收到验证失败回复（无 nome、无验证码）: {reply_text[:80]!r}")
                    log(f"[*] 5 分钟后重新发送 CPF: {cpf_text} ...")
                    await asyncio.sleep(300)
                    await send_query()
                    attempts = 0
                    log(f"[*] 已重新发送，继续等待回复...")
                    continue

                attempts += 1
                
            if is_timeout:
                log(f"    -> 等待回复超时。")
                ws.cell(row=r_idx, column=5, value="查询超时")
            else:
                text_lower = reply_text.lower()
                if "nome" in text_lower:
                    match = re.search(r'(?i)nome[\s:]+([^\n]+)', reply_text)
                    if match:
                        extracted_name = match.group(1).strip()
                        log(f"    -> 成功提取到名字: {extracted_name}")
                        ws.cell(row=r_idx, column=5, value=extracted_name)
                        
                        last_successful_reply_text = reply_text
                        last_processed_cpf = cpf_text
                    else:
                        log(f"    -> 无法用正则提取出名字。原始回复: {reply_text[:50]}...")
                        if is_retry:
                            ws.cell(row=r_idx, column=5, value="")
                        else:
                            ws.cell(row=r_idx, column=5, value="提取失败")
                elif "não encontrado" in text_lower or "nao encontrado" in text_lower or "inválido" in text_lower or "invalido" in text_lower or "utilize o comando" in text_lower or "seguido dos" in text_lower or "dígitos do cpf" in text_lower:
                    log(f"    -> CPF 未找到或无效，跳过此号码。")
                    ws.cell(row=r_idx, column=5, value="无")
                else:
                    if is_retry:
                        log(f"    -> 重试仍未成功，将结果留空。")
                        ws.cell(row=r_idx, column=5, value="")
                    else:
                        if "muitas requisições" in text_lower:
                            log(f"    -> 验证码处理未能成功通过。")
                            ws.cell(row=r_idx, column=5, value="遇到验证码且未能通过")
                        else:
                            log(f"    -> 未收到包含 nome 的回复。")
                            ws.cell(row=r_idx, column=5, value="提取失败")
                
            try:
                wb.save(excel_path)
            except PermissionError:
                log(f"[!] 保存失败：请不要在 Excel 软件中打开此表格文件！会导致文件被锁定。")
                
            await asyncio.sleep(0.2)
            return True

        total_count = 0
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=4).value and str(ws.cell(row=r, column=4).value).strip():
                total_count += 1
        
        processed_count = 0
        row = 2
        while True:
            has_more = await process_cpf_row(row, is_retry=False)
            if not has_more:
                break
            cpf_cell = ws.cell(row=row, column=4).value
            if cpf_cell and str(cpf_cell).strip():
                processed_count += 1
                progress_msg = f"[*] 进度提示：现在是 {processed_count}/{total_count} (共需处理 {total_count} 条，当前已处理 {processed_count} 条)"
                log(progress_msg)
                if progress_callback:
                    progress_callback(progress_msg)
            row += 1

        log("\n=======================================================")
        log("[*] 第一遍查询完成！现在开始重新查询之前失败的号码...")
        log("=======================================================")
        
        retry_rows = []
        for r in range(2, ws.max_row + 1):
            status_val = str(ws.cell(row=r, column=5).value or "").strip()
            if status_val in ["遇到验证码且未能通过", "查询超时", "提取失败"]:
                retry_rows.append(r)
        
        total_retry = len(retry_rows)
        retry_processed = 0
        
        row = 2
        while True:
            has_more = await process_cpf_row(row, is_retry=True)
            if not has_more:
                log(f"[*] 失败重试环节处理完成！")
                break
            if row in retry_rows:
                retry_processed += 1
                progress_msg = f"[*] 重试进度提示：现在是 {retry_processed}/{total_retry} (重试总数: {total_retry} 条，当前重试到第 {retry_processed} 条)"
                log(progress_msg)
                if progress_callback:
                    progress_callback(progress_msg)
            row += 1

        log("[*] 所有任务执行完毕！")
        await context.close()

if __name__ == '__main__':
    asyncio.run(run_cpf_query())
