import asyncio
import os
from openpyxl import load_workbook
from playwright.async_api import async_playwright

LOGIN_URL = "https://accounts.dsers.com/accounts/login"

async def run_dsers_rename(excel_path: str, user_data_dir: str, headless: bool = False, progress_callback=None):
    import builtins
    def _print(*args, **kwargs):
        builtins.print(*args, **kwargs)
        if progress_callback:
            progress_callback(" ".join(str(a) for a in args))
            
    print = _print

    print(f"[*] 加载 Excel 文件: {excel_path}")
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"[!] 无法加载 Excel 文件: {e}")
        return

    async with async_playwright() as p:
        print("[*] 启动浏览器...")
        video_dir = os.path.join(os.getcwd(), "videos")
        os.makedirs(video_dir, exist_ok=True)
        
        context = await p.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            channel="chrome",  # 使用本地 Chrome
            headless=headless,    
            viewport={'width': 1280, 'height': 800},
            record_video_dir=video_dir,  # 录制视频
            record_video_size={'width': 1280, 'height': 800}
        )
        
        page = await context.new_page()
        
        # 1. 访问并处理登录
        print(f"[*] 访问 {LOGIN_URL} ...")
        await page.goto(LOGIN_URL, wait_until='domcontentloaded', timeout=120000)
        
        print("[*] 正在等待进入系统。如果您看到登录界面，请手动输入账号密码并点击登录（或等待系统自动跳过）...")
        
        # 智能等待：一直检查直到当前页面不再是登录页
        while "login" in page.url.lower():
            await asyncio.sleep(1)
            
        print("[*] 检测到已脱离登录页面。等待页面加载缓冲...")
        await asyncio.sleep(3)
        
        # 2. 直接访问指定的 AliExpress 页面 (绕过点击侧边栏)，如果跳转不成功再尝试点击
        target_url = "https://www.dsers.com/application/orders/159831080"
        print(f"[*] 尝试直接导航至订单页: {target_url}")
        try:
            if target_url not in page.url:
                await page.goto(target_url, wait_until='domcontentloaded', timeout=60000)
            await asyncio.sleep(5)
        except Exception as e:
            print(f"[!] 直接导航失败，尝试点击左侧菜单: {e}")
            try:
                await page.click("text=Open Orders", timeout=10000)
                await page.click("text=AliExpress", timeout=10000)
                await asyncio.sleep(5)
            except Exception as e2:
                print(f"[!] 左侧菜单点击也失败: {e2}")

        search_icon_selector = "svg[data-icon='search'], i.anticon-search, .dsers-icon-search, button[aria-label='Search']"
        try:
            await page.wait_for_selector(search_icon_selector, state="visible", timeout=30000)
        except Exception:
            print("[!] 未找到搜索图标。请检查页面是否正确加载至 AliExpress 订单页。")

        print("[*] 准备就绪，开始处理表格数据...")
        
        # 3. 循环处理数据
        row = 2
        while True:
            order_id = ws.cell(row=row, column=1).value  # A列：搜索单号
            new_name = ws.cell(row=row, column=5).value  # E列：新名字
            
            if not order_id:
                print(f"[*] 第 {row} 行遇到空单号，处理结束！")
                break
                
            status_cell = ws.cell(row=row, column=6).value  # F列：状态记录
            if status_cell and str(status_cell).strip() != "":
                print(f"[*] 第 {row} 行已有处理结果 ({status_cell})，跳过...")
                row += 1
                continue
                
            order_id = str(order_id).strip()
            new_name = str(new_name).strip() if new_name else ""
            
            print(f"\n[{row}] 开始处理单号: {order_id} -> 计划修改为: '{new_name}'")
            
            # --- 步骤 3.1: 搜索 ---
            try:
                search_icon_selector = "img.searchSign[src*='search.png']"
                input_selector = "div.index_flexHeaderItem__GdUPX.searchSign input.ant-input, input.ant-input"
                
                try:
                    await page.wait_for_selector(search_icon_selector, state="visible", timeout=5000)
                    await page.click(search_icon_selector, force=True)
                    print("  -> 已点击搜索图标(放大镜)，等待搜索框弹出...")
                    await page.wait_for_selector(input_selector, state="visible", timeout=5000)
                except Exception as e:
                    pass

                await page.wait_for_selector(input_selector, state="visible", timeout=10000)
                target_input = page.locator(input_selector).first
                
                await target_input.fill("")
                await target_input.fill(order_id)
                
                ok_btn_selector = "button.ant-searchinput-btn"
                ok_btn = await page.query_selector(ok_btn_selector)
                if ok_btn and await ok_btn.is_visible():
                    await ok_btn.click(force=True)
                    print("  -> 已点击 OK 按钮执行搜索，等待页面刷新...")
                else:
                    await target_input.press("Enter")
                    print("  -> 已按回车执行搜索操作，等待页面刷新...")
                
                try:
                    await page.wait_for_selector(".ant-spin-spinning", state="visible", timeout=1000)
                except Exception:
                    pass
                try:
                    await page.wait_for_selector(".ant-spin-spinning", state="hidden", timeout=15000)
                except Exception:
                    pass
                await asyncio.sleep(1.0)
            except Exception as e:
                print(f"  [!] 搜索操作失败: {e}")
                ws.cell(row=row, column=6, value="搜索操作失败")
                wb.save(excel_path)
                row += 1
                continue
            
            # --- 步骤 3.2: 遍历分类栏寻找订单 ---
            found_category = False
            try:
                more_menu_selectors = [".ant-tabs-nav-more", "[aria-label='more']", "span:has-text('...')"]
                for more_sel in more_menu_selectors:
                    more_btn = await page.query_selector(more_sel)
                    if more_btn and await more_btn.is_visible():
                        print("  -> 发现省略号菜单，尝试展开以显示全部栏目...")
                        await more_btn.hover()
                        await asyncio.sleep(1) 
                        break 

                tab_selector = ".ant-tabs-tab, [role='tab'], .dsers-tabs-tab, .ant-dropdown-menu-item"
                tabs = await page.query_selector_all(tab_selector)
                
                for tab in tabs:
                    if await tab.is_visible():
                        text = await tab.inner_text()
                        if "(1)" in text:
                            print(f"  -> 找到匹配订单分类: {text.strip()}")
                            await tab.click()
                            found_category = True
                            try:
                                await page.wait_for_selector("text=Customer Detail", state="visible", timeout=10000)
                            except Exception:
                                await asyncio.sleep(1)
                            break
            except Exception as e:
                print(f"  [!] 查找分类标签时出错: {e}")

            if not found_category:
                print(f"  -> 找不到包含该单号的分类栏。")
                ws.cell(row=row, column=6, value="link")
                wb.save(excel_path)
                row += 1
                continue
            
            # --- 步骤 3.3: 展开订单详情 ---
            try:
                is_already_open = await page.evaluate('''() => {
                    let dialogs = Array.from(document.querySelectorAll('.ant-drawer-content, .ant-modal-content, [role="dialog"]'));
                    let container = dialogs.length > 0 ? dialogs[dialogs.length - 1] : document.body;
                    let inps = Array.from(container.querySelectorAll('input[type="text"], input:not([type])'));
                    for (let inp of inps) {
                        let rect = inp.getBoundingClientRect();
                        let cls = (inp.className || "").toString();
                        if (rect.width > 0 && rect.height > 0 && !cls.includes('search')) {
                            let parent = inp.parentElement;
                            let textContext = "";
                            let depth = 0;
                            while (parent && parent !== container && depth < 6) {
                                textContext += " " + parent.innerText;
                                parent = parent.parentElement;
                                depth++;
                            }
                            if (textContext.includes("Contact Name") || textContext.includes("Name")) {
                                return true;
                            }
                        }
                    }
                    return false;
                }''')

                if is_already_open:
                    print("  -> Customer Detail 已处于展开状态，跳过点击。")
                else:
                    await page.wait_for_selector("text=Customer Detail", state="visible", timeout=10000)
                    detail_btns = await page.locator("text=Customer Detail").all()
                    clicked = False
                    for btn in detail_btns:
                        if await btn.is_visible():
                            await btn.click()
                            clicked = True
                            break
                    
                    if clicked:
                        print("  -> 已点击 Customer Detail")
                        try:
                            await page.wait_for_selector(".ant-drawer-content, .ant-modal-content, [role='dialog']", state="visible", timeout=5000)
                            await page.wait_for_selector(".ant-drawer-content input, .ant-modal-content input, [role='dialog'] input", state="visible", timeout=5000)
                        except Exception:
                            await asyncio.sleep(1)
                    else:
                        raise Exception("Customer Detail 按钮均不可见")
            except Exception as e:
                print(f"  [!] 找不到或无法点击 Customer Detail: {e}")
                ws.cell(row=row, column=6, value="找不到详情入口")
                wb.save(excel_path)
                row += 1
                continue
                
            # --- 步骤 3.4: 修改 Contact Name ---
            try:
                target_found = await page.evaluate('''() => {
                    let dialogs = Array.from(document.querySelectorAll('.ant-drawer-content, .ant-modal-content, [role="dialog"]'));
                    let container = dialogs.length > 0 ? dialogs[dialogs.length - 1] : document.body;
                    
                    let inps = Array.from(container.querySelectorAll('input[type="text"], input:not([type])'));
                    let target = null;
                    
                    for (let i = 0; i < inps.length; i++) {
                        let inp = inps[i];
                        let rect = inp.getBoundingClientRect();
                        if (rect.width > 0 && rect.height > 0) {
                            let parent = inp.parentElement;
                            let textContext = "";
                            let depth = 0;
                            while (parent && parent !== container && depth < 6) {
                                textContext += " " + parent.innerText;
                                parent = parent.parentElement;
                                depth++;
                            }
                            if (textContext.includes("Contact Name") || textContext.includes("Name")) {
                                target = inp;
                                break;
                            }
                        }
                    }
                    
                    if (!target) {
                        for (let i = 0; i < inps.length; i++) {
                            let inp = inps[i];
                            let rect = inp.getBoundingClientRect();
                            if (rect.width > 0 && rect.height > 0 && inp.value && !inp.className.includes('search')) {
                                target = inp;
                                break;
                            }
                        }
                    }
                    
                    if (target) {
                        target.setAttribute('data-target-input', 'true');
                        return target.value;
                    }
                    return null;
                }''')
                
                if target_found is not None:
                    current_name = target_found
                    print(f"  -> 提取当前名字: '{current_name}'")
                    
                    target_input = page.locator('input[data-target-input="true"]')
                    
                    if current_name == new_name:
                        print("  -> 新旧名字一致，无需修改。")
                        ws.cell(row=row, column=6, value="成功(未修改)")
                    else:
                        print(f"  -> 执行修改: '{current_name}' -> '{new_name}'")
                        
                        await target_input.fill("")
                        await target_input.fill(new_name)
                        
                        # --- 步骤 3.5: 保存 ---
                        save_btn = page.locator("button:has-text('Save'), .ant-btn-primary:has-text('Save')").last
                        if await save_btn.is_visible():
                            if await save_btn.is_enabled():
                                await save_btn.click()
                                print("  -> 已点击 Save 按钮。")
                            else:
                                print("  [!] Save 按钮依然是灰色，尝试点击空白处让输入框失焦...")
                                await page.mouse.click(0, 0)
                                await page.wait_for_timeout(500)
                                if await save_btn.is_enabled():
                                    await save_btn.click()
                                    print("  -> 失焦后点击 Save 按钮。")
                        else:
                            print("  [!] 找不到可见的 Save 按钮！")
                            
                        # 修复: 取消未完成的任务防止 Timeout 崩溃
                        try:
                            save_wait_task1 = asyncio.create_task(page.wait_for_selector(".ant-message-notice-content", state="visible", timeout=4000))
                            save_wait_task2 = asyncio.create_task(page.wait_for_selector(".ant-drawer-content, .ant-modal-content", state="hidden", timeout=4000))
                            done, pending = await asyncio.wait([save_wait_task1, save_wait_task2], return_when=asyncio.FIRST_COMPLETED)
                            for task in pending:
                                task.cancel()
                        except Exception:
                            await asyncio.sleep(1)
                            
                        ws.cell(row=row, column=6, value="成功")
                        
                        try:
                            await target_input.evaluate("el => el.removeAttribute('data-target-input')")
                        except Exception:
                            pass
                else:
                    print("  [!] 在弹窗中找不到 Contact Name 输入框！")
                    ws.cell(row=row, column=6, value="找不到输入框")

            except Exception as e:
                 print(f"  [!] 修改名字过程出错: {e}")
                 ws.cell(row=row, column=6, value="修改出错")
                 
            try:
                wb.save(excel_path)
                close_btn = await page.query_selector("button[aria-label='Close']")
                if close_btn and await close_btn.is_visible():
                    await close_btn.click()
            except PermissionError:
                print(f"[!] 保存 Excel 失败：请不要在 Excel 软件中打开此表格文件！会导致文件被锁定。")
            except Exception:
                pass
                
            row += 1

        print("[*] 所有任务执行完毕！")
        await context.close()

if __name__ == '__main__':
    pass
